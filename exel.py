from openpyxl import load_workbook
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet")
# wb = load_workbook("기증주인수증.xlsx")
# ws = wb.active

# title = ws.cell(column=2, row=10).value

# print(title)